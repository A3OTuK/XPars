import asyncio
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
import re
from urllib.parse import unquote, urlparse, parse_qs
from openpyxl import Workbook, load_workbook
import os

CHANNELS_FILE = 'CHANNELS_LIST.txt'
RESULTS_EXCEL = 'TELEGRAM_LINKS.xlsx'
MAX_CONCURRENT_TASKS = 5  # Оптимальный баланс между скоростью и нагрузкой
DRIVER_TIMEOUT = 10  # Таймаут для WebDriverWait
PAGE_LOAD_DELAY = 2  # Базовая задержка для загрузки страницы

# Глобальный счетчик обработанных ссылок
processed_count = 0


def init_excel_file():
    """Инициализация Excel-файла с заголовком"""
    if not os.path.exists(RESULTS_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Telegram Links"
        ws['A1'] = "Telegram Links"
        wb.save(RESULTS_EXCEL)
        print(f"✅ Создан новый файл {RESULTS_EXCEL}")


async def save_to_excel(telegram_link):
    """Асинхронное добавление ссылки в Excel-файл"""
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _save_link_sync, telegram_link)

        # Обновляем счетчик после успешного сохранения
        global processed_count
        processed_count += 1
    except Exception as e:
        print(f"❌ Ошибка при сохранении в Excel: {e}")


def _save_link_sync(telegram_link):
    """Синхронная часть сохранения в Excel"""
    try:
        wb = load_workbook(RESULTS_EXCEL)
        ws = wb.active
        row = 1
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        ws.cell(row=row, column=1, value=telegram_link)
        wb.save(RESULTS_EXCEL)
    except Exception as e:
        raise e


async def fetch_telegram_link(driver, channel_url):
    try:
        await asyncio.get_event_loop().run_in_executor(None, driver.get, channel_url)
        await asyncio.sleep(PAGE_LOAD_DELAY)

        # Переход на вкладку "О канале"
        try:
            about_button = WebDriverWait(driver, DRIVER_TIMEOUT).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[contains(text(), "О канале") or contains(text(), "About")]')))
            about_button.click()
            await asyncio.sleep(PAGE_LOAD_DELAY / 2)
        except (NoSuchElementException, TimeoutException):
            pass

        telegram_links = await _find_redirect_links(driver)
        if not telegram_links:
            telegram_links = await _find_description_links(driver)

        return telegram_links[0] if telegram_links else None
    except Exception as e:
        print(f"⚠️ Ошибка при обработке {channel_url}: {str(e)}")
        return None


async def _find_redirect_links(driver):
    """Поиск ссылок через редирект YouTube"""
    try:
        links = WebDriverWait(driver, DRIVER_TIMEOUT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a[href*="redirect"]')))

        telegram_links = []
        for link in links:
            href = link.get_attribute("href")
            if "youtube.com/redirect" in href:
                parsed_url = urlparse(href)
                query_params = parse_qs(parsed_url.query)
                if "q" in query_params:
                    decoded_url = unquote(query_params["q"][0])
                    if "t.me" in decoded_url or "telegram.me" in decoded_url:
                        telegram_links.append(decoded_url)
        return telegram_links
    except TimeoutException:
        return []


async def _find_description_links(driver):
    """Поиск ссылок в описании канала"""
    try:
        description = WebDriverWait(driver, DRIVER_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#description, .description')))

        matches = re.finditer(r'(?:@|t\.me\/)([a-zA-Z0-9_]{5,})', description.text)
        return [f"https://t.me/{m.group(1)}" for m in matches]
    except TimeoutException:
        return []


async def process_single_channel(channel, semaphore, driver_pool):
    async with semaphore:
        driver = None
        try:
            # Используем драйвер из пула
            driver = await driver_pool.get_driver()
            print(f"\n🔍 Обрабатываю канал: {channel}")

            telegram_link = await fetch_telegram_link(driver, channel)

            if telegram_link:
                clean_link = re.sub(r'https?://(www\.)?(telegram\.me|t\.me)', 'https://t.me', telegram_link)
                clean_link = clean_link.split('?')[0].rstrip('/')
                print(f"✅ Найдена ссылка: {clean_link} (Всего найдено: {processed_count + 1})")
                await save_to_excel(clean_link)
        except Exception as e:
            print(f"⚠️ Ошибка при обработке {channel}: {str(e)}")
        finally:
            if driver:
                await driver_pool.release_driver(driver)


class DriverPool:
    """Пул драйверов для повторного использования"""

    def __init__(self, max_drivers):
        self.max_drivers = max_drivers
        self.drivers = []
        self.semaphore = asyncio.Semaphore(max_drivers)

    async def get_driver(self):
        await self.semaphore.acquire()
        if self.drivers:
            return self.drivers.pop()
        else:
            options = webdriver.ChromeOptions()
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--mute-audio")
            return webdriver.Chrome(options=options)

    async def release_driver(self, driver):
        self.drivers.append(driver)
        self.semaphore.release()

    async def close_all(self):
        for driver in self.drivers:
            driver.quit()
        self.drivers = []


async def process_channels():
    init_excel_file()

    try:
        with open(CHANNELS_FILE, 'r', encoding='utf-8') as f:
            channels = [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"❌ Файл {CHANNELS_FILE} не найден!")
        return

    print(f"\n📋 Найдено {len(channels)} каналов для обработки")

    driver_pool = DriverPool(MAX_CONCURRENT_TASKS)
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_TASKS)

    try:
        tasks = [process_single_channel(channel, semaphore, driver_pool) for channel in channels]
        await asyncio.gather(*tasks)
    finally:
        await driver_pool.close_all()

    print(f"\n🏁 Обработка завершена! Найдено {processed_count} ссылок. Результаты в {RESULTS_EXCEL}")


if __name__ == "__main__":
    if not os.path.exists(CHANNELS_FILE):
        print(f"❌ Ошибка: Файл {CHANNELS_FILE} не найден!")
    else:
        print("🚀 Запуск оптимизированного парсера Telegram-ссылок")
        asyncio.run(process_channels())

        if os.path.exists(RESULTS_EXCEL):
            print(f"✔️ Файл результатов создан: {os.path.abspath(RESULTS_EXCEL)}")